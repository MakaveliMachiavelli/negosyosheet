#!/usr/bin/env python3
"""NegosyoSheet generator — builds the full + demo xlsx products.
Usage: python3 build_sheet.py full   → dist/NegosyoSheet.xlsx
       python3 build_sheet.py demo   → dist/NegosyoSheet-Demo.xlsx (fewer rows unlocked, watermark sheet)
"""
import sys, datetime, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

MODE = sys.argv[1] if len(sys.argv) > 1 else "full"
DIST = os.path.join(os.path.dirname(__file__), "dist")
os.makedirs(DIST, exist_ok=True)

# ---------- palette ----------
NAVY = "0B5FFF"; DARK = "101828"; GREY = "667085"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color=DARK)
LINK_FONT = Font(color=NAVY, underline="single")
THIN = Border(*[Side(style="thin", color="E4E7EC")]*4)
MONEY = '#,##0.00'
DATEF = 'yyyy-mm-dd'

wb = Workbook()

def style_header(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

# ================= INSTRUCTIONS =================
ws = wb.active; ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100
rows = [
    ("title", "NEGOSYOSHEET — Inventory, Utang at Profit Tracker"),
    ("sub", "Para sa sari-sari store, online reseller, at maliliit na negosyo"),
    ("blank", ""),
    ("h", "PAANO GAMITIN (3 steps lang):"),
    ("li", "1. I-fill ang INVENTORY sheet — item name, cost, selling price, stock-in, at reorder point."),
    ("li", "2. I-log ang bawat benta sa SALES LOG (may dropdown na item list, auto-compute ang profit)."),
    ("li", "3. I-log ang utang sa UTANG at mga gastos sa EXPENSES. Watch ang DASHBOARD — lahat nandoon."),
    ("blank", ""),
    ("h", "MGA SHEET:"),
    ("li", "• DASHBOARD — total sales, profit, inventory value, utang, at reorder alerts. Dito ang summary."),
    ("li", "• INVENTORY — listahan ng items. Auto-compute: current stock, stock value, REORDER NA alert."),
    ("li", "• SALES LOG — bawat sale. Auto: unit price (galing Inventory), total, cost, profit."),
    ("li", "• UTANG — listahan ng mga utang. Auto: days overdue at KULIKAHIN NA! alert."),
    ("li", "• EXPENSES — kuryente, upa, load, at iba. Ibaon sa DASHBOARD computation."),
    ("blank", ""),
    ("h", "IMPORTANTE:"),
    ("li", "• Wag i-edit ang mga may formula (kulay grey). Type lang sa white cells."),
    ("li", "• Gumagana sa Excel at Google Sheets (i-upload lang sa Drive → Open with Google Sheets)."),
    ("li", "• Ang profit sa SALES LOG = total benta − cost ng items. Ang NET sa DASHBOARD = profit − expenses."),
    ("blank", ""),
    ("sub2", f"NegosyoSheet v1.0 · {datetime.date.today().isoformat()} · Support: check README sa page where you downloaded"),
]
r = 2
for kind, text in rows:
    cell = ws.cell(row=r, column=2, value=text)
    if kind == "title": cell.font = TITLE_FONT
    elif kind == "sub": cell.font = Font(size=12, color=GREY)
    elif kind == "sub2": cell.font = Font(size=9, color=GREY)
    elif kind == "h": cell.font = Font(bold=True, size=12, color=DARK)
    elif kind == "li": cell.font = Font(size=11)
    r += 1
if MODE == "demo":
    ws.cell(row=r+1, column=2, value="*** DEMO VERSION — limited to 5 inventory rows. Get the full version para walang limit! ***").font = Font(bold=True, color="DC2626")

N_ROWS = 5 if MODE == "demo" else 200  # data rows per sheet

# ================= INVENTORY =================
ws = wb.create_sheet("INVENTORY")
headers = ["Item Name", "Unit", "Cost (₱)", "Selling Price (₱)", "Stock In", "Stock Out", "Current Stock", "Stock Value (₱)", "Reorder Point", "Status"]
for i, h in enumerate(headers, 1): ws.cell(row=1, column=i, value=h)
style_header(ws, 1, range(1, len(headers)+1))
widths = [26, 8, 12, 16, 10, 11, 13, 15, 13, 16]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
sample = [("Noka Noodles", "pc", 12, 18, 48, 12, 12), ("Softdrinks 1L", "btl", 68, 90, 24, 5, 6), ("Sabon 60g", "pc", 22, 35, 30, 8, 10)]
for j, s in enumerate(sample if MODE == "demo" else sample, start=2):
    for i, v in enumerate(s, 1): ws.cell(row=j, column=i, value=v)
for r in range(2, N_ROWS + 2):
    ws.cell(row=r, column=7, value=f"=IF(A{r}=\"\",\"\",E{r}-F{r})")
    ws.cell(row=r, column=8, value=f"=IF(A{r}=\"\",\"\",G{r}*C{r})")
    ws.cell(row=r, column=10, value=f'=IF(A{r}="","",IF(G{r}<=I{r},"⚠️ REORDER NA!","OK"))')
    for c in range(1, 11):
        cell = ws.cell(row=r, column=c); cell.border = THIN
        if c in (3, 4, 8): cell.number_format = MONEY
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{N_ROWS+1}"
ws.conditional_formatting.add(f"J2:J{N_ROWS+1}",
    FormulaRule(formula=[f'$J2="⚠️ REORDER NA!"'], fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color="DC2626", bold=True)))
ws.conditional_formatting.add(f"J2:J{N_ROWS+1}",
    FormulaRule(formula=[f'$J2="OK"'], fill=PatternFill("solid", fgColor="D1F0E1"), font=Font(color="067A53")))

# ================= SALES LOG =================
ws = wb.create_sheet("SALES LOG")
headers = ["Date", "Item Name", "Qty", "Unit Price (₱)", "Total (₱)", "Cost (₱)", "Profit (₱)"]
for i, h in enumerate(headers, 1): ws.cell(row=1, column=i, value=h)
style_header(ws, 1, range(1, len(headers)+1))
for i, w in enumerate([12, 26, 8, 14, 13, 12, 13], 1): ws.column_dimensions[get_column_letter(i)].width = w
today = datetime.date.today().isoformat()
for r in range(2, N_ROWS + 2):
    ws.cell(row=r, column=4, value=f'=IF(OR(B{r}="",C{r}=""),"",IFERROR(VLOOKUP(B{r},INVENTORY!A:D,4,FALSE),""))')
    ws.cell(row=r, column=5, value=f'=IF(OR(B{r}="",C{r}=""),"",D{r}*C{r})')
    ws.cell(row=r, column=6, value=f'=IF(OR(B{r}="",C{r}=""),"",IFERROR(VLOOKUP(B{r},INVENTORY!A:C,3,FALSE),"")*C{r})')
    ws.cell(row=r, column=7, value=f'=IF(OR(B{r}="",C{r}=""),"",E{r}-F{r})')
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c); cell.border = THIN
        if c in (4, 5, 6, 7): cell.number_format = MONEY
        if c == 1: cell.number_format = DATEF
for j, (d, it, q) in enumerate([(today, "Noka Noodles", 3), (today, "Softdrinks 1L", 1)], start=2):
    ws.cell(row=j, column=1, value=d); ws.cell(row=j, column=2, value=it); ws.cell(row=j, column=3, value=q)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{N_ROWS+1}"
dv = DataValidation(type="list", formula1=f"=INVENTORY!$A$2:$A${N_ROWS+1}", allow_blank=True, showDropDown=False)
ws.add_data_validation(dv); dv.add(f"B2:B{N_ROWS+1}")
ws.conditional_formatting.add(f"G2:G{N_ROWS+1}",
    CellIsRule(operator="lessThan", formula=["0"], font=Font(color="DC2626", bold=True)))

# ================= UTANG =================
ws = wb.create_sheet("UTANG")
headers = ["Date", "Customer", "Item / Notes", "Amount (₱)", "Due Date", "Paid?", "Days Overdue", "Status"]
for i, h in enumerate(headers, 1): ws.cell(row=1, column=i, value=h)
style_header(ws, 1, range(1, len(headers)+1))
for i, w in enumerate([12, 20, 24, 13, 12, 9, 12, 16], 1): ws.column_dimensions[get_column_letter(i)].width = w
for r in range(2, N_ROWS + 2):
    ws.cell(row=r, column=7, value=f'=IF(OR(E{r}="",F{r}="Yes"),"",MAX(0,TODAY()-E{r}))')
    ws.cell(row=r, column=8, value=f'=IF(F{r}="Yes","✅ PAID",IF(A{r}="","",IF(G{r}>0,"🔴 KULIKAHIN NA!","🕒 Hindi pa due")))')
    for c in range(1, 9):
        cell = ws.cell(row=r, column=c); cell.border = THIN
        if c == 4: cell.number_format = MONEY
        if c in (1, 5): cell.number_format = DATEF
dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv); dv.add(f"F2:F{N_ROWS+1}")
ws.conditional_formatting.add(f"H2:H{N_ROWS+1}",
    FormulaRule(formula=['$H2="🔴 KULIKAHIN NA!"'], fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color="DC2626", bold=True)))
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:H{N_ROWS+1}"

# ================= EXPENSES =================
ws = wb.create_sheet("EXPENSES")
headers = ["Date", "Category", "Description", "Amount (₱)"]
for i, h in enumerate(headers, 1): ws.cell(row=1, column=i, value=h)
style_header(ws, 1, range(1, len(headers)+1))
for i, w in enumerate([12, 18, 30, 13], 1): ws.column_dimensions[get_column_letter(i)].width = w
dv = DataValidation(type="list", formula1='"Kuryente,Upa,Tubig,Load/Internet,Delivery/Transport,Supplies,Iba pa"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv); dv.add(f"B2:B{N_ROWS+1}")
for r in range(2, N_ROWS + 2):
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c); cell.border = THIN
        if c == 4: cell.number_format = MONEY
        if c == 1: cell.number_format = DATEF
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:D{N_ROWS+1}"

# ================= DASHBOARD =================
ws = wb.create_sheet("DASHBOARD", 0)  # first sheet
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for col, w in zip("BCDEFG", [30, 18, 4, 30, 18, 4]): ws.column_dimensions[col].width = w
last = N_ROWS + 1
ws.cell(row=2, column=2, value="📊 DASHBOARD").font = TITLE_FONT
ws.cell(row=3, column=2, value="Auto-compute lahat — wag i-edit ang mga numbers dito").font = Font(color=GREY, size=10)

def card(ws, row, col, label, formula, fmt='#,##0.00', fill="F6F8FB"):
    ws.cell(row=row, column=col, value=label).font = Font(bold=True, size=11, color=GREY)
    c = ws.cell(row=row+1, column=col, value=formula)
    c.font = Font(bold=True, size=18, color=DARK); c.number_format = fmt
    for cc in range(col, col+2):
        for rr in range(row-1 if False else row, row+2):
            pass
    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill)
    ws.cell(row=row+1, column=col).fill = PatternFill("solid", fgColor=fill)
    ws.cell(row=row, column=col+1).fill = PatternFill("solid", fgColor=fill)
    ws.cell(row=row+1, column=col+1).fill = PatternFill("solid", fgColor=fill)

card(ws, 5, 2, "TOTAL SALES (₱)", f"=SUM('SALES LOG'!E2:E{last})")
card(ws, 5, 5, "TOTAL GROSS PROFIT (₱)", f"=SUM('SALES LOG'!G2:G{last})")
card(ws, 8, 2, "TOTAL EXPENSES (₱)", f"=SUM(EXPENSES!D2:D{last})")
card(ws, 8, 5, "NET (profit − expenses) (₱)", f"=SUM('SALES LOG'!G2:G{last})-SUM(EXPENSES!D2:D{last})")
card(ws, 11, 2, "INVENTORY VALUE (₱)", f"=SUM(INVENTORY!H2:H{last})")
card(ws, 11, 5, "UTANG NA HINDI PAID (₱)", f'=SUMIFS(UTANG!D2:D{last},UTANG!F2:F{last},"<>Yes",UTANG!A2:A{last},"<>")')
card(ws, 14, 2, "ITEMS NA REORDER NA", f'=COUNTIF(INVENTORY!J2:J{last},"⚠️ REORDER NA!")', fmt='0', fill="FEF7E6")
card(ws, 14, 5, "MGA UTANG NA LATE", f'=COUNTIF(UTANG!H2:H{last},"🔴 KULIKAHIN NA!")', fmt='0', fill="FEF7E6")

ws.cell(row=17, column=2, value="TOTAL PER ITEM (sales):").font = Font(bold=True, size=11)
ws.cell(row=17, column=5, value="TOP BUYERS (utang):").font = Font(bold=True, size=11)
ws.cell(row=18, column=2, value="Item").font = Font(bold=True, size=9, color=GREY)
ws.cell(row=18, column=3, value="Sales (₱)").font = Font(bold=True, size=9, color=GREY)
for r in range(2, min(12, N_ROWS+1)):
    rr = 17 + r - 1
    ws.cell(row=rr+1, column=2, value=f"=IF(INVENTORY!A{r}=\"\",\"\",INVENTORY!A{r})").font = Font(size=10)
    c = ws.cell(row=rr+1, column=3, value=f"=IF(INVENTORY!A{r}=\"\",\"\",SUMIF('SALES LOG'!B:B,INVENTORY!A{r},'SALES LOG'!E:E))")
    c.number_format = MONEY; c.font = Font(size=10)

if MODE == "demo":
    ws.cell(row=30, column=2, value="DEMO VERSION — 5 rows lang ang pwede. Full version: unlimited rows + free updates.").font = Font(bold=True, color="DC2626")

out = os.path.join(DIST, "NegosyoSheet.xlsx" if MODE == "full" else "NegosyoSheet-Demo.xlsx")
wb.save(out)
print("saved", out, os.path.getsize(out), "bytes")
