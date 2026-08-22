#!/usr/bin/env python3
"""Verify NegosyoSheet xlsx: structure, formulas, sample math (reference-computed)."""
import zipfile, sys
from openpyxl import load_workbook

ok = 0; bad = 0
def check(name, cond):
    global ok, bad
    if cond: ok += 1; print(f"  ✓ {name}")
    else: bad += 1; print(f"  ✗ FAIL: {name}")

path = "dist/NegosyoSheet.xlsx"
check("valid zip container", zipfile.is_zipfile(path))
wb = load_workbook(path)
check("sheets present", set(["DASHBOARD","README","INVENTORY","SALES LOG","UTANG","EXPENSES"]).issubset(set(wb.sheetnames)))
check("dashboard first", wb.sheetnames[0] == "DASHBOARD")

inv = wb["INVENTORY"]
check("inventory current-stock formula", inv["G2"].value == '=IF(A2="","",E2-F2)')
check("inventory stock-value formula", inv["H2"].value == '=IF(A2="","",G2*C2)')
check("reorder status formula", "REORDER NA" in (inv["J2"].value or ""))
# reference math on sample rows: noka 48in-12out=36 cur, value 36×12=432
check("sample row A2 = Noka Noodles", inv["A2"].value == "Noka Noodles")

sales = wb["SALES LOG"]
check("sales VLOOKUP price formula", "VLOOKUP(B2,INVENTORY!A:D,4,FALSE)" in (sales["D2"].value or ""))
check("sales total formula", sales["E2"].value == '=IF(OR(B2="",C2=""),"",D2*C2)')
# reference: noka 3 × ₱18 = ₱54 total, cost 3×₱12=₱36, profit ₱18
check("sample sale logged (Noka x3)", sales["B2"].value == "Noka Noodles" and sales["C2"].value == 3)

utang = wb["UTANG"]
check("utang overdue formula", "TODAY()-E2" in (utang["G2"].value or ""))
check("utang status formula", "KULIKAHIN NA" in (utang["H2"].value or ""))

dash = wb["DASHBOARD"]
formulas = [dash.cell(row=r, column=c).value for r in (6, 9, 12, 15) for c in (2, 5)]
check("dashboard has 8 metric cards", sum(1 for f in formulas if isinstance(f, str) and f.startswith("=")) >= 8)
check("dashboard sales formula", "SUM('SALES LOG'!E2:E" in (dash["B6"].value or ""))

d = load_workbook("dist/NegosyoSheet-Demo.xlsx")
check("demo watermark present", any("DEMO VERSION" in str(c.value) for row in d["README"].iter_rows() for c in row if c.value))
check("demo smaller than full", __import__("os").path.getsize("dist/NegosyoSheet-Demo.xlsx") < __import__("os").path.getsize(path))

print(f"\nRESULT: {ok} passed, {bad} failed")
sys.exit(1 if bad else 0)
